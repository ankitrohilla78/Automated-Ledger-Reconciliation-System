import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ==============================
# CONFIG
# ==============================
file_path = "C:/MCard/Repayment_Automation/Repayment_Automation.xlsx"

PAID_EMI_MAP = {
    "Cred EMI": 4
}

# ==============================
# LOAD DATA
# ==============================
df = pd.read_excel(file_path)
df.columns = df.columns.str.strip()

df["EMI Date"] = pd.to_datetime(df["EMI Date"], errors="coerce", dayfirst=True)

today = datetime.today()

# ==============================
# STATUS LOGIC
# ==============================
def get_status(row):
    if pd.isna(row["EMI Date"]):
        return "Unknown"

    loan = row["Loan Source"]

    if loan in PAID_EMI_MAP:
        if row["Month"] <= PAID_EMI_MAP[loan]:
            return "Paid"
        elif row["EMI Date"] <= today:
            return "Overdue"
        else:
            return "Upcoming"
    else:
        if row["EMI Date"] < today:
            return "Paid"
        elif row["EMI Date"].date() == today.date():
            return "Due Today"
        else:
            return "Upcoming"

df["Status"] = df.apply(get_status, axis=1)

# ==============================
# LOAD WORKBOOK
# ==============================
wb = load_workbook(file_path)
ws = wb.active

headers = [cell.value for cell in ws[1]]

if "Status" not in headers:
    status_col = len(headers) + 1
    ws.cell(row=1, column=status_col, value="Status")
else:
    status_col = headers.index("Status") + 1

# ==============================
# COLORS
# ==============================
green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

# ==============================
# APPLY STATUS + COLORS
# ==============================
for i, row in df.iterrows():
    excel_row = i + 2
    status = row["Status"]

    ws.cell(row=excel_row, column=status_col, value=status)

    if status == "Paid":
        fill = green
    elif status == "Upcoming":
        fill = yellow
    elif status == "Overdue":
        fill = red
    elif status == "Due Today":
        fill = orange
    else:
        fill = None

    if fill:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=excel_row, column=col).fill = fill

# ==============================
# SUMMARY SHEET
# ==============================
df_paid = df[df["EMI Date"] <= today]

df_latest = (
    df_paid.sort_values("EMI Date")
    .groupby("Loan Source")
    .last()
    .reset_index()
)

# ---- REMAINING MONTHS ----
total_months = df.groupby("Loan Source")["Month"].max().to_dict()

def calc_remaining(row):
    return total_months[row["Loan Source"]] - row["Month"]

df_latest["Remaining Months"] = df_latest.apply(calc_remaining, axis=1)

# ---- INTEREST CALCULATION ----
# Total interest per loan
total_interest_map = df.groupby("Loan Source")["Interest"].sum().to_dict()

# Paid interest per loan
paid_interest_map = df_paid.groupby("Loan Source")["Interest"].sum().to_dict()

def calc_pending_interest(row):
    loan = row["Loan Source"]
    total_interest = total_interest_map.get(loan, 0)
    paid_interest = paid_interest_map.get(loan, 0)
    return total_interest - paid_interest

df_latest["Total Interest Pending"] = df_latest.apply(calc_pending_interest, axis=1)

# Remove old sheet
if "Latest_Updates" in wb.sheetnames:
    wb.remove(wb["Latest_Updates"])

# Create new sheet
ws_new = wb.create_sheet("Latest_Updates")

# Write headers
ws_new.append(df_latest.columns.tolist())

# Write rows
for row in df_latest.itertuples(index=False):
    ws_new.append(row)

# ==============================
# SAVE
# ==============================
wb.save(file_path)

print("✅ DONE: Interest Pending + Remaining Months + Full System Ready")